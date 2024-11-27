import asyncio
import hashlib
import uuid
from dataclasses import dataclass, field
from enum import IntEnum

from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.worksheet import Worksheet
from concurrent.futures import ThreadPoolExecutor


@dataclass
class BaseMenu:
    id: uuid.UUID
    title: str
    description: str


@dataclass
class Menu(BaseMenu):
    pass


@dataclass
class Submenu(BaseMenu):
    menu_id: uuid.UUID


@dataclass
class Dish(BaseMenu):
    price: float
    discount: None | float = None
    submenu_id: uuid.UUID = None

    def __post_init__(self) -> None:
        self.discount = 0 if self.discount is None else self.discount


@dataclass
class RestaurantMenu:
    menus: list[Menu] = field(default_factory=list)
    submenus: list[Submenu] = field(default_factory=list)
    dishes: list[Dish] = field(default_factory=list)


class ColumnMenu(IntEnum):
    ID = 1
    TITLE = 2
    DESCRIPTION = 3


class ColumnSubmenu(IntEnum):
    ID = 2
    TITLE = 3
    DESCRIPTION = 4


class ColumnDish(IntEnum):
    ID = 3
    TITLE = 4
    DESCRIPTION = 5
    PRICE = 6
    DISCOUNT = 7


class ParserXlsxService:
    def __init__(self) -> None:
        self.sheet:  Worksheet | None = None
        self.wb: Workbook | None = None
        self.__path: str | None = None
        self.__hash: str | None = None
        self.__hash_values = set()

    def load_sheet(self, path: str) -> None:
        self.__path = path
        self.wb = load_workbook(filename=path)
        self.sheet = self.wb.active

    def construct_entity(self,
                         entity_type: type[BaseMenu],
                         row: int,
                         column_type: type[IntEnum],
                         entity_id: uuid.UUID = None) -> BaseMenu | None:
        values = [self.sheet.cell(row=row, column=column).value for column in column_type]
        hash_values = sum(hash(value) for value in values[1:])
        if (len(values) < 5 and not all(values)) or hash_values in self.__hash_values:
            return
        self.__hash_values.add(hash_values)
        values[0] = uuid.uuid4()
        if entity_id:
            values.append(entity_id)
        return entity_type(*values)

    async def get_restaurant_menu(self) -> RestaurantMenu | None:
        if not await self._check_hash_sheet():
            return
        restaurant_menu = RestaurantMenu()
        rows = (_ for _ in range(1, self.sheet.max_row + 1))
        menu_id, submenu_id = None, None
        for row in rows:
            if menu := self.construct_entity(Menu, row, ColumnMenu):
                menu_id = menu.id
                restaurant_menu.menus.append(menu)
                row = next(rows)
            if submenu := self.construct_entity(Submenu, row, ColumnSubmenu, menu_id):
                submenu_id = submenu.id
                restaurant_menu.submenus.append(submenu)
                row = next(rows)
            if dish := self.construct_entity(Dish, row, ColumnDish, submenu_id):
                restaurant_menu.dishes.append(dish)
        return restaurant_menu

    def _generate_hash(self) -> str:
        hash_ = hashlib.sha256()
        with open(self.__path, 'rb') as file:
            hash_.update(file.read())
        return hash_.hexdigest()

    async def _check_hash_sheet(self) -> bool:
        with ThreadPoolExecutor() as executor:
            loop = asyncio.get_running_loop()
            hash_ = await loop.run_in_executor(executor, self._generate_hash)
        if self.__hash is None or self.__hash != hash_:
            self.__hash = hash_
            return True
        return False


if __name__ == '__main__':
    async def main():
        parser = ParserXlsxService()
        parser.load_sheet('../admin/Menu_2.xlsx')
        restaurant_menu = await parser.get_restaurant_menu()
        assert restaurant_menu is not None, "restaurant_menu should be not None"
        for menu in restaurant_menu.menus:
            print(menu)
        for submenu in restaurant_menu.submenus:
            print(submenu)
        for dish in restaurant_menu.dishes:
            print(dish)
        restaurant_menu_2 = await parser.get_restaurant_menu()
        assert restaurant_menu_2 is None, "restaurant_menu should None"

    asyncio.run(main())
