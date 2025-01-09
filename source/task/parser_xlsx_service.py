import hashlib
import uuid
from dataclasses import dataclass, field
from enum import IntEnum

from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from pydantic import BaseModel

from database.schemas import MenuCreation, SubmenuCreation, DishCreation


@dataclass
class RestaurantMenu:
    menu_id_to_menu: dict[str, MenuCreation] = field(default_factory=dict)
    menu_id_submenu_id_to_submenu: dict[tuple[str, ...], SubmenuCreation] = field(default_factory=dict)
    menu_id_submenu_id_dish_id_to_dish: dict[tuple[str, ...], DishCreation] = field(default_factory=dict)


class ColumnMenu(IntEnum):
    ID = 1
    TITLE = 2
    DESCRIPTION = 3


class ColumnSubmenu(IntEnum):
    ID = 2
    TITLE = 3
    DESCRIPTION = 4


class ColumnDish(IntEnum):
    ID = 3
    TITLE = 4
    DESCRIPTION = 5
    PRICE = 6
    DISCOUNT = 7


class ParserXlsxService:
    def __init__(self) -> None:
        self.sheet:  Worksheet | None = None
        self.book: Workbook | None = None
        self.path: str | None = None
        self.hash_file: str | None = None

    async def load_sheet(self, path: str) -> None:
        self.path = path
        self.book = load_workbook(filename=path)
        self.sheet = self.book.active

    @staticmethod
    async def generate_hash(file_path: str, mode='rb') -> str:
        hash_ = hashlib.sha256()
        with open(file_path, mode) as file:
            hash_.update(file.read())
        return hash_.hexdigest()

    async def check_hash_file(self, file_path: str, mode='rb') -> bool:
        hash_ = await self.generate_hash(file_path, mode)
        if self.hash_file is None or self.hash_file != hash_:
            self.hash_file = hash_
            return True
        return False

    def construct_entity(self,
                         entity_type: type[BaseModel],
                         row: int,
                         column_type: type[IntEnum],
                         entity_id: uuid.UUID = None) -> BaseModel | None:
        cells = [self.sheet.cell(row=row, column=column) for column in column_type]
        values = [cell.value for cell in cells]
        if not all(values[:-1]):
            return

        if not self._check_uuid_4(values[0]):
            values[0] = uuid.uuid4()
            cells[0].value = str(values[0])
            self.book.save(self.path)
            self.hash_file = self.generate_hash(self.path)
        if entity_id is not None:
            values.append(entity_id)
        keys = entity_type.model_fields.keys()
        return entity_type(**dict(zip(keys, values)))

    def get_restaurant_menu(self) -> RestaurantMenu:
        restaurant_menu = RestaurantMenu()
        rows = iter(range(1, self.sheet.max_row + 1))
        menu_id, submenu_id = None, None
        for row in rows:
            if menu := self.construct_entity(MenuCreation, row, ColumnMenu):
                menu_id = str(menu.id)
                restaurant_menu.menu_id_to_menu[menu_id] = menu
                row = next(rows)
            if menu_id and (submenu := self.construct_entity(SubmenuCreation, row, ColumnSubmenu, menu_id)):
                submenu_id = str(submenu.id)
                restaurant_menu.menu_id_submenu_id_to_submenu[(menu_id, submenu_id)] = submenu
                row = next(rows)
            if submenu_id and (dish := self.construct_entity(DishCreation, row, ColumnDish, submenu_id)):
                restaurant_menu.menu_id_submenu_id_dish_id_to_dish[(menu_id, submenu_id, str(dish.id))] = dish
        self.sheet = None
        self.book = None
        return restaurant_menu

    @staticmethod
    def _check_uuid_4(value: str) -> bool:
        try:
            uuid.UUID(value, version=4)
            return True
        except ValueError:
            return False
